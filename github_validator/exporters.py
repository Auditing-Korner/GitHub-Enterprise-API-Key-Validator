"""
Export Module

Provides export functionality for reports in various formats:
- PDF export
- Excel export
- JSON export (enhanced)
"""

from typing import Dict, List, Optional, Any
import json
import os
from datetime import datetime


class ReportExporter:
    """Exports reports to various formats."""
    
    def __init__(self):
        self.supported_formats = ["json", "pdf", "excel", "csv"]
    
    def export_to_json(self, data: Dict[str, Any], output_file: str) -> str:
        """
        Export data to JSON format.
        
        Args:
            data: Data to export
            output_file: Output file path
            
        Returns:
            Path to exported file
        """
        if not output_file.endswith('.json'):
            output_file += '.json'
        
        with open(output_file, 'w', encoding='utf-8') as f:
            json.dump(data, f, indent=2, ensure_ascii=False, default=str)
        
        return output_file
    
    def export_to_pdf(self, html_content: str, output_file: str) -> str:
        """
        Export HTML report to PDF format.
        
        Args:
            html_content: HTML content of the report
            output_file: Output file path
            
        Returns:
            Path to exported file
        """
        try:
            import weasyprint
        except ImportError:
            raise ImportError(
                "weasyprint is required for PDF export. Install it with: pip install weasyprint"
            )
        
        if not output_file.endswith('.pdf'):
            output_file += '.pdf'
        
        weasyprint.HTML(string=html_content).write_pdf(output_file)
        return output_file
    
    def export_to_excel(self, data: Dict[str, Any], output_file: str) -> str:
        """
        Export data to Excel format.
        
        Args:
            data: Data to export
            output_file: Output file path
            
        Returns:
            Path to exported file
        """
        try:
            import pandas as pd
            from openpyxl import Workbook
            from openpyxl.styles import Font, PatternFill, Alignment
        except ImportError:
            raise ImportError(
                "pandas and openpyxl are required for Excel export. Install with: pip install pandas openpyxl"
            )
        
        if not output_file.endswith('.xlsx'):
            output_file += '.xlsx'
        
        wb = Workbook()
        wb.remove(wb.active)  # Remove default sheet
        
        # Export permissions
        if "permissions" in data:
            self._add_permissions_sheet(wb, data["permissions"])
        
        # Export enumeration
        if "enumeration" in data:
            self._add_enumeration_sheet(wb, data["enumeration"])
        
        # Export resources
        if "resources" in data:
            self._add_resources_sheet(wb, data["resources"])
        
        wb.save(output_file)
        return output_file
    
    def _add_permissions_sheet(self, wb, permissions_data: Dict[str, Any]):
        """Add permissions data to Excel workbook."""
        from openpyxl.styles import Font, PatternFill, Alignment
        
        ws = wb.create_sheet("Permissions")
        
        # Headers
        headers = ["Permission", "Status", "Message", "Details"]
        ws.append(headers)
        
        # Style headers
        header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF")
        
        for cell in ws[1]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center")
        
        # Add critical permissions
        critical_perms = permissions_data.get("critical_permissions", {})
        for perm_name, perm_data in critical_perms.items():
            status = "GRANTED" if perm_data.get("granted") else "DENIED"
            ws.append([
                perm_name,
                status,
                perm_data.get("message", ""),
                json.dumps(perm_data.get("details", {}))
            ])
        
        # Add standard permissions
        standard_perms = permissions_data.get("standard_permissions", {})
        for perm_name, perm_data in standard_perms.items():
            status = "GRANTED" if perm_data.get("granted") else "DENIED"
            ws.append([
                perm_name,
                status,
                perm_data.get("message", ""),
                json.dumps(perm_data.get("details", {}))
            ])
        
        # Auto-adjust column widths
        for column in ws.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)
            ws.column_dimensions[column_letter].width = adjusted_width
    
    def _add_enumeration_sheet(self, wb, enumeration_data: Dict[str, Any]):
        """Add enumeration data to Excel workbook."""
        ws = wb.create_sheet("Enumeration")
        
        # Add organization info
        if "organization_info" in enumeration_data:
            ws.append(["Organization Information"])
            org_info = enumeration_data["organization_info"]
            for key, value in org_info.items():
                ws.append([key, str(value)])
            ws.append([])  # Empty row
        
        # Add members
        if "members" in enumeration_data:
            ws.append(["Members"])
            ws.append(["Login", "ID", "Type", "Site Admin"])
            for member in enumeration_data["members"]:
                ws.append([
                    member.get("login", ""),
                    member.get("id", ""),
                    member.get("type", ""),
                    member.get("site_admin", False)
                ])
            ws.append([])
        
        # Add repositories
        if "repositories" in enumeration_data:
            ws.append(["Repositories"])
            ws.append(["Name", "Full Name", "Private", "Fork", "Stars", "Forks"])
            for repo in enumeration_data["repositories"][:100]:  # Limit to 100
                ws.append([
                    repo.get("name", ""),
                    repo.get("full_name", ""),
                    repo.get("private", False),
                    repo.get("fork", False),
                    repo.get("stargazers_count", 0),
                    repo.get("forks_count", 0)
                ])
    
    def _add_resources_sheet(self, wb, resources_data: Dict[str, Any]):
        """Add resources data to Excel workbook."""
        ws = wb.create_sheet("Resources")
        
        # Add repositories
        if "repositories" in resources_data:
            repos = resources_data["repositories"]
            if isinstance(repos, dict) and "repositories" in repos:
                ws.append(["Repositories"])
                ws.append(["Full Name", "Private", "Fork"])
                for repo in repos["repositories"][:100]:
                    ws.append([
                        repo.get("full_name", ""),
                        repo.get("private", False),
                        repo.get("fork", False)
                    ])
        
        # Add secrets
        if "secrets" in resources_data:
            secrets = resources_data["secrets"]
            if isinstance(secrets, list):
                ws.append([])
                ws.append(["Organization Secrets"])
                ws.append(["Name", "Created At", "Updated At"])
                for secret in secrets[:50]:
                    ws.append([
                        secret.get("name", ""),
                        secret.get("created_at", ""),
                        secret.get("updated_at", "")
                    ])


def export_report(report_data: Dict[str, Any], output_file: str, format: str = "json") -> str:
    """
    Export report to specified format.
    
    Args:
        report_data: Report data dictionary
        output_file: Output file path
        format: Export format (json, pdf, excel, csv)
        
    Returns:
        Path to exported file
    """
    exporter = ReportExporter()
    
    if format.lower() == "json":
        return exporter.export_to_json(report_data, output_file)
    elif format.lower() == "pdf":
        if "html_content" not in report_data:
            raise ValueError("HTML content required for PDF export")
        return exporter.export_to_pdf(report_data["html_content"], output_file)
    elif format.lower() in ["excel", "xlsx"]:
        return exporter.export_to_excel(report_data, output_file)
    else:
        raise ValueError(f"Unsupported format: {format}. Supported formats: {exporter.supported_formats}")

